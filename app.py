#!/usr/bin/env python3.7

# Copyright (C) Aryason Consulting - All Rights Reserved
# Unauthorized copying of this project, via any medium is strictly prohibited
# Proprietary and confidential

from flask import Flask, request, make_response, jsonify
from flask_restful import Api, Resource
from functools import wraps
import json
from json import JSONDecodeError
import os
from flask_cors import CORS
from excel_generator import generate_excel_file
from datetime import datetime
import database
import openpyxl
from math import sin, cos, sqrt, atan2, radians

UPLOAD_FOLDER = 'technical_visits'

app = Flask(__name__)
CORS(app)
api = Api(app)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

type_materiau_general = ['S355', 'S235']
type_materiau_boulon_tige = ['4.6', '4.8', '5.6', '5.8', '6.8', '8.8', '10.9']


def auth_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        auth = request.authorization
        if auth and database.exist(auth.username, auth.password):
            return f(*args, **kwargs)
        return make_response('Authentication failed', 401, {'WWW-Authenticate': 'Basic realm="login Required"'})

    return decorated


class Authentication(Resource):

    @auth_required
    def get(self):

        wb = openpyxl.load_workbook('./data/sites.xlsx')
        ws = wb.active

        codes_site = []

        row = 2
        while ws['A' + str(row)].value is not None:
            codes_site.append(ws['A' + str(row)].value)
            row += 1

        return codes_site


def get_troncons_number_and_height_by_code_site(code_site):
    try:
        if code_site.isnumeric():
            code_site = int(code_site)
            if code_site < 100:
                troncons_number = code_site
                height = troncons_number * 6
                return troncons_number, height

    finally:
        wb = openpyxl.load_workbook('./data/sites.xlsx')
        ws = wb.active
        row_found = False
        row = 1
        while ws['A' + str(row)].value is not None:
            if str(ws['A' + str(row)].value).strip() == str(code_site).strip():
                row_found = True
                break
            row += 1
        if row_found:
            if ws['Z' + str(row)].value is not None:
                height = float(ws['Z' + str(row)].value)
            else:
                height = 0
            troncons_number = round(height / 6)
        else:
            troncons_number = 1
            height = 6

    return troncons_number, height


# TODO: link with the database
def get_data(trancons_number):
    trancon = {
        "numero": "1",
        "zsup": "3.2",
        "zinf": "",
        "bsup": "20.3",
        "binf": "",
        "membrures": {
            "tube": "true",
            "longueur": "",
            "diametre": "2.2",
            "epaisseur": "3.3",
            "b": "",
            "H": "",
            "materiau": "S355"
        },
        "diagonales": {
            "tube": "false",
            "longueur": "1.2",
            "diametre": "",
            "epaisseur": "3.4",
            "b": "4.5",
            "H": "5.6",
            "materiau": ""
        },
        "traverses": {
            "tube": "true",
            "longueur": "",
            "diametre": "2.2",
            "epaisseur": "",
            "b": "",
            "H": "",
            "materiau": "S355"
        },
        "dtiges": "1.1",
        "dbride": "",
        "drepartition": "3.3",
        "ebride": "",
        "mat_tiges": "4.6",
        "mat_plaque": "S355",
        "nb_tiges": "",
        "mat_boulon": "4.6",
        "mat_bride": "S355",
        "nb_boulons": ""
    }
    detailed_form = {
        "trancons": [],
        "hms": "3.3",
        "lf": "2.2",
        "hf": "3.3",
        "h": "",
        "a": "5.6",
        "b": "",
        "elu": "",
        "els": "200.5",
        "commentaire": "test commentaire"
    }
    for i in range(trancons_number):
        trancon['numero'] = i + 1
        detailed_form['trancons'].append(trancon)

    entry = {
        "Type": "Antennes_Radio",
        "Diametre": "",
        "constructeur": "6878300G",
        "az": "",
        "hma": "18,95",
        "nb_coax": "",
        "type_coax": '7/8"',
        "commentaire": ""
    }

    simplified_form = []
    for i in range(trancons_number):
        simplified_form.append(entry)

    data = {'formulaire_detaillé': detailed_form, 'formulaire_simplifie': simplified_form}

    return data


class simple_form(Resource):

    @auth_required
    def post(self):

        # Input validation : code site
        try:
            code_site = request.form['code_site']
        except ValueError as error:
            return 'Invalid input code_site : ' + str(error), 400

        # Create folder
        dt = datetime.now()
        folder_name = str(code_site) + '_' + str(dt).replace(' ', '_').replace('.', '_').replace(':', '_')
        folder_path = os.path.join('.', UPLOAD_FOLDER, folder_name)
        os.mkdir(folder_path)

        # Create Excel file
        wb = openpyxl.load_workbook('./data/EBD.xlsx')
        ws = wb.active

        # Save Excel
        data = json.loads(request.form['data'])
        for i, row in enumerate(data):
            ws['B' + str(i + 4)].value = row['Type']  # type
            ws['K' + str(i + 4)].value = row['Diametre']  # diametre
            ws['P' + str(i + 4)].value = row['constructeur']  # constructueur
            ws['W' + str(i + 4)].value = row['az']  # az
            ws['Y' + str(i + 4)].value = row['hma']  # hma
            ws['AA' + str(i + 4)].value = row['nb_coax']  # nb coax
            ws['AC' + str(i + 4)].value = row['type_coax']  # type coax
            ws['AE' + str(i + 4)].value = row['commentaire']  # commentaire
        wb.save(folder_path + '/VT_simple.xlsx')

        # Save photos
        files = request.files
        file_argument = 'photo_simplified'
        if file_argument in files:
            file = files[file_argument]
            file_extension = file.filename.split('.')[1].lower()
            file_name = 'photo' + '.' + file_extension
            # Save file
            file.save(os.path.join(folder_path, file_name))

        return True


def distance_as_the_crow_flies(latitude_1, longitude_1, latitude_2, longitude_2):
    # Source : https://stackoverflow.com/questions/19412462/getting-distance-between-two-points-based-on-latitude-longitude

    # approximate radius of earth in km
    R = 6373.0

    latitude_1 = radians(float(latitude_1))
    longitude_1 = radians(float(longitude_1))
    latitude_2 = radians(float(latitude_2))
    longitude_2 = radians(float(longitude_2))

    dlon = longitude_2 - longitude_1
    dlat = latitude_2 - latitude_1

    a = sin(dlat / 2) ** 2 + cos(latitude_1) * cos(latitude_2) * sin(dlon / 2) ** 2
    c = 2 * atan2(sqrt(a), sqrt(1 - a))

    distance = R * c

    return distance


def get_troncons_number_and_height_by_coordinates(latitude, longitude):

    distances = []

    wb = openpyxl.load_workbook('./data/sites.xlsx')
    ws = wb.active
    row = 2
    while ws['A' + str(row)].value is not None:
        latitude_2 = ws['AD' + str(row)].value
        longitude_2 = ws['AE' + str(row)].value
        if latitude_2 is None or longitude_2 is None:
            latitude_2 = '0'
            longitude_2 = '0'
        distance = distance_as_the_crow_flies(latitude, longitude, latitude_2.replace(',', '.'), longitude_2.replace(',', '.'))
        distances.append(distance)
        row += 1

    row = 2 + distances.index(min(distances))
    code_site = ws['A' + str(row)].value

    if ws['Z' + str(row)].value is not None:
        height = float(ws['Z' + str(row)].value)
    else:
        height = 0
    troncons_number = round(height / 6)

    return troncons_number, height, code_site


def is_location_correct(code_site, longitude, latitude):
    wb = openpyxl.load_workbook('./data/sites.xlsx')
    ws = wb.active
    row = 2

    while ws['A' + str(row)].value is not None and ws['A' + str(row)].value != code_site:
        row += 1

    latitude_site = ws['AD' + str(row)].value
    longitude_site = ws['AE' + str(row)].value
    distance = distance_as_the_crow_flies(latitude, longitude, latitude_site, longitude_site)

    max_autorised_distance = 50

    return distance < max_autorised_distance


class data(Resource):

    @auth_required
    def get(self):
        args = request.args
        code_site = args.get('code_site', default=None)
        latitude = args.get('latitude', default=None)
        longitude = args.get('longitude', default=None)
        if latitude == '0' and longitude == '0':
            return 'Le code site est incorrect', 400
        if code_site is None and (latitude is None or longitude is None) :
            return 'Code site, latitude ou longitude manquant', 400
        if (latitude is None or longitude is None) and code_site is not None :
            troncons_number, height = get_troncons_number_and_height_by_code_site(code_site)
        elif code_site is None and latitude is not None and longitude is not None:
            troncons_number, height, code_site = get_troncons_number_and_height_by_coordinates(latitude, longitude)
        else:
            if not is_location_correct(code_site, longitude, latitude):
                return False
            troncons_number, height = get_troncons_number_and_height_by_code_site(code_site)

        data = get_data(troncons_number) # code_site should be also a prameter when linked to the database

        return jsonify({'nombre_troncon': troncons_number, 'hauteur': height, 'code_site':code_site, 'data': data})


def data_type(v):
    if v is None:
        return None

    # s = str(v)
    # regex = re.compile(r'(?P<int>^([\d]+)$)|(?P<float>^([\d*\.\d+])$)|(?P<bool>True|False)|(?P<string>.+)')
    # try:
    #     return regex.search(s).lastgroup
    # except AttributeError:
    #     return None

    if v == 'true' or v == 'false':
        return 'bool'

    s = str(type(v))
    s = s.replace("<class '", "").replace("'>", "")
    return 'int'


# ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'JPG'}
#
#
# def allowed_file(filename):
#     return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


class detailed_form(Resource):

    @auth_required
    def post(self):

        # Input validation : code site
        try:
            code_site = request.form['code_site']
        except ValueError as error:
            return 'Invalid input code_site : ' + str(error), 400

        # Input validation : data
        try:
            data = json.loads(request.form['data'])

        except JSONDecodeError as error:
            return 'Invalid input data : ' + str(error), 400
        try:
            for trancon in data['trancons']:
                numero = trancon['numero']
                zsup = trancon['zsup']
                bsup = trancon['bsup']
                if data_type(numero) != 'int':
                    return 'Invalid input numero', 400
                if data_type(zsup) not in ['float', 'int']:
                    return 'Invalid input zsup', 400
                if data_type(bsup) not in ['float', 'int']:
                    return 'Invalid input bsup', 400
                for tp in ['membrures', 'diagonales', 'traverses']:

                    tube = trancon[tp]['tube']
                    longueur = trancon[tp]['longueur']
                    epaisseur = trancon[tp]['epaisseur']
                    diametre = trancon[tp]['diametre']
                    materiau = trancon[tp]['materiau']
                    b = trancon[tp]['b']
                    h = trancon[tp]['H']
                    if data_type(longueur) not in ['float', 'int']:
                        return 'Invalid input ' + tp + ' longueur', 400
                    if data_type(epaisseur) not in ['float', 'int']:
                        return 'Invalid input ' + tp + ' epaisseur', 400
                    if data_type(tube) != 'bool':
                        return 'Invalid input ' + tp + ' tube', 400
                    if tube:
                        if data_type(diametre) not in ['float', 'int']:
                            return 'Invalid input ' + tp + ' diametre', 400
                    else:
                        if data_type(b) not in ['float', 'int']:
                            return 'Invalid input ' + tp + ' b', 400
                        if data_type(h) not in ['float', 'int']:
                            return 'Invalid input ' + tp + ' H', 400

                    if materiau not in type_materiau_general:
                        return 'Invalid input ' + materiau + ' materiau', 400
                if numero == 1:

                    mat_tiges = trancon['mat_tiges']
                    mat_plaque = trancon['mat_plaque']
                    nb_tiges = trancon['nb_tiges']
                    binf = trancon['binf']
                    zinf = trancon['zinf']
                    if mat_tiges not in type_materiau_boulon_tige:
                        return 'Invalid input ' + mat_tiges + ' mat_tiges', 400
                    if mat_plaque not in type_materiau_general:
                        return 'Invalid input ' + mat_plaque + ' mat_plaque', 400
                    if data_type(nb_tiges) != 'int':
                        return 'Invalid input ' + nb_tiges + ' nb_tiges', 400
                    if data_type(binf) not in ['int', 'float']:
                        return 'Invalid input ' + binf + ' binf', 400
                    if data_type(zinf) not in ['int', 'float']:
                        return 'Invalid input ' + zinf + ' zinf', 400
                else:
                    mat_boulon = trancon['mat_boulon']
                    mat_bride = trancon['mat_bride']
                    nb_boulons = trancon['nb_boulons']
                    if mat_boulon not in type_materiau_boulon_tige:
                        return 'Invalid input ' + mat_boulon + ' mat_tiges', 400
                    if mat_bride not in type_materiau_general:
                        return 'Invalid input ' + mat_bride + ' mat_plaque', 400
                    if data_type(nb_boulons) != 'int':
                        return 'Invalid input ' + nb_boulons + ' nb_tiges', 400
            hms = data['hms']
            lf = data['lf']
            hf = data['hf']
            h = data['h']
            a = data['a']
            b = data['b']
            elu = data['elu']
            els = data['els']
            commentaire = data['commentaire']
            if data_type(hms) not in ['float', 'int']:
                return 'Invalid input hms', 400
            if data_type(lf) not in ['float', 'int']:
                return 'Invalid input lf', 400
            if data_type(hf) not in ['float', 'int']:
                return 'Invalid input hf', 400
            if data_type(h) not in ['float', 'int']:
                return 'Invalid input h', 400
            if data_type(a) not in ['float', 'int']:
                return 'Invalid input a', 400
            if data_type(b) not in ['float', 'int']:
                return 'Invalid input b', 400
            if data_type(elu) not in ['float', 'int']:
                return 'Invalid input elu', 400
            if data_type(els) not in ['float', 'int']:
                return 'Invalid input elu', 400
        except KeyError as error:
            return 'Invalid input Key Error : ' + str(error), 400

        # Check that the sent number of trancons is correct
        trancons_number, _ = get_troncons_number_and_height_by_code_site(code_site)
        gotten_trancons_number = len(data['trancons'])
        if gotten_trancons_number != trancons_number:
            return 'Gotten trancons number is incorrect', 400

        dt = datetime.now()
        folder_name = str(code_site) + '_' + str(dt).replace(' ', '_').replace('.', '_').replace(':', '_')
        folder_path = os.path.join('.', UPLOAD_FOLDER, folder_name)
        os.mkdir(folder_path)

        # Input validation : photos
        # https://roytuts.com/python-flask-rest-api-file-upload/
        for i in range(1, trancons_number + 1):

            for element in ['troncon', 'membrures', 'diagonales', 'traverses', 'bride']:

                file_argument = element + '_' + str(i)
                if file_argument in request.files:
                    file = request.files[file_argument]
                    if file:
                        # Create file name
                        file_extension = file.filename.split('.')[1].lower()
                        file_name = element + '_' + str(i) + '.' + file_extension
                        # Save file
                        file.save(os.path.join(folder_path, file_name))

        for element in ['site', 'base']:
            if element in request.files:

                file = request.files[element]
                if file:
                    # Create file name
                    file_extension = file.filename.split('.')[1].lower()
                    file_name = element + '.' + file_extension
                    # Save file
                    file.save(os.path.join(folder_path, file_name))

        generate_excel_file(code_site_=code_site, data_=data, file_name_=os.path.join(folder_path, 'VT.xlsx'))

        return True


api.add_resource(Authentication, '/authentication')
api.add_resource(data, '/data')
api.add_resource(simple_form, '/simple_form')
api.add_resource(detailed_form, '/detailed_form')

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
