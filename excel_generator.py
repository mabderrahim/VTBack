from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment


def trancon_by_numero(data_, numero):

    i = 0
    trancon = data_['trancons'][i]
    while trancon['numero'] != numero and i < len(data_['trancons']):
        trancon = data_['trancons'][i]
        i += 1
    if trancon['numero'] != numero:
        raise Exception("No trancon found with numero " + str(numero))

    return trancon


def section(element):

    if element['tube'] == 'true':
        return 'CHS-' + str(float(element['diametre']) * float(element['epaisseur']))
    else:
        return 'LS-' + str(float(element['b']) * float(element['H']) * float(element['epaisseur']))


def style_1(cell=None, size=14):
    cell.fill = PatternFill(fill_type="solid", start_color='000000', end_color='000000')
    cell.font = Font(bold=True, name="Arial", size=size, color="FFFFFFFF")
    cell.alignment = Alignment(horizontal="center", vertical="center")


def style_2(cells):
    for cell in cells:
        cell.fill = PatternFill(fill_type="solid", start_color='AE28C3', end_color='AE28C3')
        cell.font = Font(bold=True, name="Arial", size=10, color="FFFFFFFF")


def style_3(cells):
    for cell in cells:
        cell.fill = PatternFill(fill_type="solid", start_color='F2CFF3', end_color='F2CFF3')
        cell.font = Font(bold=True, name="Arial", size=10, color='000000')


def input_style(cell):
    cell.font = Font(bold=True, color="0000FF")


def generate_excel_file(code_site_=None, data_=None, file_name_=None):

    # Styles tutorial :
    # https://www.blog.pythonlibrary.org/2021/08/11/styling-excel-cells-with-openpyxl-and-python/

    wb = Workbook()
    ws = wb.active

    # Columns width
    ws.column_dimensions['A'].width = 11.86 + 0.72 + 1
    ws.column_dimensions['B'].width = 13.57 + 0.72
    ws.column_dimensions['C'].width = 11.86 + 0.72
    ws.column_dimensions['D'].width = 11.86 + 0.72
    ws.column_dimensions['E'].width = 11.86 + 0.72 + 5
    ws.column_dimensions['F'].width = 11.86 + 0.72
    ws.column_dimensions['G'].width = 11.86 + 0.72
    ws.column_dimensions['H'].width = 11.86 + 0.72

    ws['A1'] = 'Visite Technique'
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=8)
    style_1(ws['A1'])

    ws['A4'] = 'Description'
    style_2([ws[i+'4'] for i in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']])

    ws['A6'] = 'Site'
    ws['B6'] = code_site_
    input_style(ws['B6'])

    ws.merge_cells(start_row=7, start_column=2, end_row=8, end_column=8)
    ws['A7'] = 'Commentaires'
    ws['B7'] = data_['commentaire']
    ws['B7'].alignment = Alignment(vertical="top")
    input_style(ws['B7'])

    ws['A10'] = 'Structure'
    ws.merge_cells(start_row=10, start_column=1, end_row=10, end_column=8)
    style_1(ws['A10'], size=12)

    ws['A12'] = 'Géométrie & Sections'
    style_2([ws[i + '12'] for i in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']])

    ws['A14'] = 'Géométrie des tronçons'
    style_3([ws[i + '14'] for i in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']])

    ws['A16'] = 'N°'
    ws['B16'] = 'Binf(m)'
    ws['C16'] = 'Bsup(m)'
    ws['D16'] = 'Zinf(m)'
    ws['E16'] = 'Zsup(m)'
    row = 16

    # Insert values of N°, Binf(m),	Bsup(m), Zinf(m), Zsup(m)
    for numero in range(1, len(data_['trancons'])+1):

        # Get trancon number 'numero'
        trancon = trancon_by_numero(data_, numero)

        row += 1

        # Insert numero
        ws['A' + str(row)] = numero

        # Insert Binf & Zinf
        if numero == 1:
            ws['B' + str(row)] = trancon['binf']
            input_style(ws['B' + str(row)])
            ws['D' + str(row)] = trancon['zinf']
            input_style(ws['D' + str(row)])

        else:
            # Insert Binf = Bsup of previous trancon
            ws['B' + str(row)] = ws['C' + str(row-1)].value
            input_style(ws['B' + str(row)])
            # Insert Zinf = Zsup of previous trancon
            ws['D' + str(row)] = ws['E' + str(row-1)].value
            input_style(ws['D' + str(row)])

        # Insert Bsup and Zsup
        ws['C' + str(row)] = trancon['bsup']
        input_style(ws['C' + str(row)])
        ws['E' + str(row)] = trancon['zsup']
        input_style(ws['E' + str(row)])

    row += 2
    ws['A' + str(row)] = 'Sections'
    style_3([ws[i + str(row)] for i in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']])

    row += 2
    ws['A' + str(row)] = 'Informations'
    ws['B' + str(row)] = "Les dimensions des sections doivent être indiquées: exemple: CHS-diamètre(en mm)xépaisseur (en mm)\npour les tubes ou LS-hauteur(en mm)xlargeur(en mm) épaisseur(en mm) pour les cornières "
    ws.row_dimensions[row].height = 35.25
    ws['A' + str(row)].alignment = Alignment(vertical="top")
    ws['B' + str(row)].alignment = Alignment(vertical="top", wrapText=True)
    ws['B' + str(row)].font = Font(name="Arial", size=10)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)

    row += 2
    ws['B' + str(row)] = 'Membrures'
    ws['D' + str(row)] = 'Diagonales'
    ws['F' + str(row)] = 'Traverses'
    ws['B' + str(row)].alignment = Alignment(horizontal="center")
    ws['D' + str(row)].alignment = Alignment(horizontal="center")
    ws['F' + str(row)].alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=5)
    ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=7)

    row += 1
    ws['A' + str(row)] = 'N°'
    ws['B' + str(row)] = 'Section'
    ws['C' + str(row)] = 'Matériau'
    ws['D' + str(row)] = 'Section'
    ws['E' + str(row)] = 'Matériau'
    ws['F' + str(row)] = 'Section'
    ws['G' + str(row)] = 'Matériau'

    ws['B' + str(row)].alignment = Alignment(horizontal="center")
    ws['C' + str(row)].alignment = Alignment(horizontal="center")
    ws['D' + str(row)].alignment = Alignment(horizontal="center")
    ws['E' + str(row)].alignment = Alignment(horizontal="center")
    ws['F' + str(row)].alignment = Alignment(horizontal="center")
    ws['G' + str(row)].alignment = Alignment(horizontal="center")

    # Insert values of N°, Section, Matériau
    for numero in range(1, len(data_['trancons'])+1):

        # Get trancon number 'numero'
        trancon = trancon_by_numero(data_, numero)

        row += 1

        # Insert numero
        ws['A' + str(row)] = numero

        # Insert Section, Matériau
        ws['B' + str(row)] = section(trancon['membrures'])
        input_style(ws['B' + str(row)])
        ws['C' + str(row)] = trancon['membrures']['materiau']
        input_style(ws['C' + str(row)])
        ws['D' + str(row)] = section(trancon['diagonales'])
        input_style(ws['D' + str(row)])
        ws['E' + str(row)] = trancon['diagonales']['materiau']
        input_style(ws['E' + str(row)])
        ws['F' + str(row)] = section(trancon['traverses'])
        input_style(ws['F' + str(row)])
        ws['G' + str(row)] = trancon['traverses']['materiau']
        input_style(ws['G' + str(row)])

    row += 2
    ws['A' + str(row)] = 'Ancrages & Assemblages'
    style_2([ws[i + str(row)] for i in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']])

    row += 2
    ws['A' + str(row)] = 'Tronçons'
    style_3([ws[i + str(row)] for i in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']])

    row += 1
    ws['B' + str(row)] = 'Plaque'
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    ws['B' + str(row)].alignment = Alignment(horizontal="center")
    ws['E' + str(row)] = 'Tiges'
    ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=8)
    ws['E' + str(row)].alignment = Alignment(horizontal="center")

    row += 1
    ws['A' + str(row)] = 'Ancrages'
    ws['B' + str(row)] = 'DBride (mm)'
    ws['C' + str(row)] = 'eBride (mm)'
    ws['D' + str(row)] = 'Matériau'
    ws['E' + str(row)] = 'DRépartition (mm)'
    ws['F' + str(row)] = 'NbTiges'
    ws['G' + str(row)] = 'DTiges (mm)'
    ws['H' + str(row)] = 'Matériau'

    row += 1
    trancon = trancon_by_numero(data_, 1)
    ws['A' + str(row)] = 0
    ws['B' + str(row)] = trancon['dbride']
    input_style(ws['B' + str(row)])
    ws['C' + str(row)] = trancon['ebride']
    input_style(ws['C' + str(row)])
    ws['D' + str(row)] = trancon['mat_plaque']
    input_style(ws['D' + str(row)])
    ws['E' + str(row)] = trancon['drepartition']
    input_style(ws['E' + str(row)])
    ws['F' + str(row)] = trancon['nb_tiges']
    input_style(ws['F' + str(row)])
    ws['G' + str(row)] = trancon['dtiges']
    input_style(ws['G' + str(row)])
    ws['H' + str(row)] = trancon['mat_tiges']
    input_style(ws['H' + str(row)])

    row += 2
    ws['B' + str(row)] = 'Bride'
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    ws['B' + str(row)].alignment = Alignment(horizontal="center")
    ws['E' + str(row)] = 'Boulons'
    ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=8)
    ws['E' + str(row)].alignment = Alignment(horizontal="center")

    row += 1
    ws['A' + str(row)] = 'Assemblages'
    ws['B' + str(row)] = 'DBride (mm)'
    ws['C' + str(row)] = 'eBride (mm)'
    ws['D' + str(row)] = 'Matériau'
    ws['E' + str(row)] = 'DRépartition (mm)'
    ws['F' + str(row)] = 'NbTiges'
    ws['G' + str(row)] = 'DTiges (mm)'
    ws['H' + str(row)] = 'Matériau'

    # Insert assemblage information
    for numero in range(2, len(data_['trancons'])+1):

        # Get trancon number 'numero'
        trancon = trancon_by_numero(data_, numero)

        row += 1

        # Insert numero
        ws['A' + str(row)] = numero - 1

        # Insert remaining information
        ws['B' + str(row)] = trancon['dbride']
        input_style(ws['B' + str(row)])
        ws['C' + str(row)] = trancon['ebride']
        input_style(ws['C' + str(row)])
        ws['D' + str(row)] = trancon['mat_bride']
        input_style(ws['D' + str(row)])
        ws['E' + str(row)] = trancon['drepartition']
        input_style(ws['E' + str(row)])
        ws['F' + str(row)] = trancon['nb_boulons']
        input_style(ws['F' + str(row)])
        ws['G' + str(row)] = trancon['dtiges']
        input_style(ws['G' + str(row)])
        ws['H' + str(row)] = trancon['mat_boulon']
        input_style(ws['H' + str(row)])

    row += 2
    ws['A' + str(row)] = 'Fondations'
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    style_1(ws['A' + str(row)], size=12)

    row += 2
    ws['A' + str(row)] = 'Si Fondations Superficielles'
    style_2([ws[i + str(row)] for i in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']])

    row += 2
    ws['A' + str(row)] = 'Description du massif'
    style_3([ws[i + str(row)] for i in ['A', 'B', 'C']])

    row += 1
    ws['A' + str(row)] = 'A'
    ws['B' + str(row)] = data_['a']
    input_style(ws['B' + str(row)])
    ws['C' + str(row)] = 'm'
    ws['D' + str(row)] = 'Longueur du Massif'

    row += 1
    ws['A' + str(row)] = 'B'
    ws['B' + str(row)] = data_['b']
    input_style(ws['B' + str(row)])
    ws['C' + str(row)] = 'm'
    ws['D' + str(row)] = 'Largeur du Massif'

    row += 1
    ws['A' + str(row)] = 'H'
    ws['B' + str(row)] = data_['h']
    input_style(ws['B' + str(row)])
    ws['C' + str(row)] = 'm'
    ws['D' + str(row)] = 'Hauteur de béton'

    row += 1
    ws['A' + str(row)] = 'Lf'
    ws['B' + str(row)] = data_['lf']
    input_style(ws['B' + str(row)])
    ws['C' + str(row)] = 'm'
    ws['D' + str(row)] = 'Largeur des plots béton'

    row += 1
    ws['A' + str(row)] = 'Hf'
    ws['B' + str(row)] = data_['hf']
    input_style(ws['B' + str(row)])
    ws['C' + str(row)] = 'm'
    ws['D' + str(row)] = 'Hauteur des plots béton dans le sol'

    row += 1
    ws['A' + str(row)] = 'Hms'
    ws['B' + str(row)] = data_['hms']
    input_style(ws['B' + str(row)])
    ws['C' + str(row)] = 'm'
    ws['D' + str(row)] = 'Hauteur des plots béton hors sol'

    row += 2
    ws['A' + str(row)] = 'Description du sol'
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    style_3([ws[i + str(row)] for i in ['A', 'B', 'C', 'D', 'E']])

    row += 1
    ws['A' + str(row)] = "Contrainte limite du sol à l'ELU (KPa):"
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    ws['D' + str(row)] = data_['elu']
    input_style(ws['D' + str(row)])

    row += 1
    ws['A' + str(row)] = "Contrainte limite du sol à l'ELS (KPa):"
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    ws['D' + str(row)] = data_['els']
    input_style(ws['D' + str(row)])

    # Save the file
    wb.save(file_name_)


# if __name__ == '__main__':
#     data = {
#         "trancons": [
#             {
#                 "numero": 1,
#                 "zsup": 10,
#                 "zinf": 30,
#                 "bsup": 20,
#                 "binf": 60,
#                 "membrures": {
#                     "tube": True,
#                     "longueur": 1,
#                     "diametre": 2,
#                     "epaisseur": 3,
#                     "b": 4,
#                     "H": 5,
#                     "materiau": "S355"
#                 },
#                 "diagonales": {
#                     "tube": False,
#                     "longueur": 1,
#                     "diametre": 2,
#                     "epaisseur": 3,
#                     "b": 4,
#                     "H": 5,
#                     "materiau": "S355"
#                 },
#                 "traverses": {
#                     "tube": True,
#                     "longueur": 1,
#                     "diametre": 2,
#                     "epaisseur": 3,
#                     "b": 4,
#                     "H": 5,
#                     "materiau": "S355"
#                 },
#                 "dtiges": 1,
#                 "dbride": 2,
#                 "drepartition": 3,
#                 "ebride": 4,
#
#                 "mat_tiges": 4.6,
#                 "mat_plaque": "S355",
#                 "nb_tiges": 5,
#
#                 "mat_boulon": 4.6,
#                 "mat_bride": "S355",
#                 "nb_boulons": 15
#             },
#             {
#                 "numero": 2,
#                 "zsup": 10,
#                 "zinf": 30,
#                 "bsup": 20,
#                 "binf": 60,
#                 "membrures": {
#                     "tube": False,
#                     "longueur": 1,
#                     "diametre": 2,
#                     "epaisseur": 3,
#                     "b": 4,
#                     "H": 5,
#                     "materiau": "S355"
#                 },
#                 "diagonales": {
#                     "tube": False,
#                     "longueur": 1,
#                     "diametre": 2,
#                     "epaisseur": 3,
#                     "b": 4,
#                     "H": 5,
#                     "materiau": "S355"
#                 },
#                 "traverses": {
#                     "tube": True,
#                     "longueur": 1,
#                     "diametre": 2,
#                     "epaisseur": 3,
#                     "b": 4,
#                     "H": 5,
#                     "materiau": "S355"
#                 },
#                 "dtiges": 1,
#                 "dbride": 2,
#                 "drepartition": 3,
#                 "ebride": 4,
#
#                 "mat_tiges": 4.6,
#                 "mat_plaque": "S355",
#                 "nb_tiges": 5,
#
#                 "mat_boulon": 4.6,
#                 "mat_bride": "S355",
#                 "nb_boulons": 15
#             }
#         ],
#         "hms": 3,
#         "lf": 2,
#         "hf": 3,
#         "h": 4,
#         "a": 5,
#         "b": 6,
#         "elu": 300,
#         "els": 200,
#         "commentaire": "test commentaire"
#     }
#     code_site = 'FR-83-900010'
#
#     generate_excel_file(code_site_=code_site, data_=data)
